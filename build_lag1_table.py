"""
build_lag1_table.py — Lag-1 forecast vs Actuals table (Jan–May 2026).

Columns:
  Material_Number | Country_Name
  Lag1_Jan_2026 | Lag1_Feb_2026 | Lag1_Mar_2026 | Lag1_Apr_2026 | Lag1_May_2026
  Actual_Jan_2026 | Actual_Feb_2026 | Actual_Mar_2026 | Actual_Apr_2026 | Actual_May_2026

Lag-1 definition: for forecast month M, the forecast made in month M-1.
Actuals sourced from adjfc_nz.parquet (freshly fetched Jun 22).
"""

import math
import os
import sys
import time

import pandas as pd
import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from aera_auth import ensure_token, login as _aera_login

DIR = os.path.dirname(os.path.abspath(__file__))

BASE_URL = "https://becleproximo.aeratechnology.com"
DATA_URL = (
    f"{BASE_URL}/ispring/awc?v=3"
    "&processID=6C9EBAEB_0F03_4A5D_AF19_7188A3AEA9C7"
    "&ServiceName=ExecuteBIObjectData"
)

BIOID   = "B4D7A7D4_94EC_4DED_8A0F_CE4291E8CA7A"
SHEETID = "AEAC976F-1BE2-453F-83F6-33428B763610"
FID     = "683F21F6_1495_4835_8474_1E09C8BDFB24"

SNAPSHOT_DIM = "F2C3C017-1EDF-11ED-A548-0A617A24E20D_4DC610B6-B7C1-4D4C-A1C6-22E9FE11F662"
FORECAST_DIM = "F2C3C017-1EDF-11ED-A548-0A617A24E20D_035317EF-A2C0-415E-B864-0F032A347371"

ROW = (
    "060CACC0-EC7A-4946-A601-C86E4D69AB29_DAB42D7F-6407-4CA3-ADA4-92456C940A47|,"  # Country Name
    "50191AB8-1EDB-11ED-A548-0A617A24E20D_9EB3B832-5F1A-4FF1-814B-CC82933E9F14|,"  # Customer Number
    "04E2EDB1-1A36-11ED-A548-0A617A24E20D_454245B2-6AF3-49B8-AA8E-18FEC4E340DC|"   # Material Number
)
MEA = "E3524740-47C7-4C30-A381-333FC13DEBD6|SUM|||||"  # Adjusted FC only

PAGE_LIMIT = 2000

# Lag pairs: (snapshot_month, forecast_month)
LAG1_PAIRS = [
    ("Dec 2025", "Jan 2026"),
    ("Jan 2026", "Feb 2026"),
    ("Feb 2026", "Mar 2026"),
    ("Mar 2026", "Apr 2026"),
    ("Apr 2026", "May 2026"),
]

LAG3_PAIRS = [
    ("Oct 2025", "Jan 2026"),
    ("Nov 2025", "Feb 2026"),
    ("Dec 2025", "Mar 2026"),
    ("Jan 2026", "Apr 2026"),
    ("Feb 2026", "May 2026"),
]

FORECAST_MONTHS = ["Jan 2026", "Feb 2026", "Mar 2026", "Apr 2026", "May 2026"]


# ── HTTP ──────────────────────────────────────────────────────────────────────

def _post(token, jsessionid, payload, lb="", retries=8):
    def _c(tok, sid, lb): c = {"JSESSIONID": sid, "accessToken": tok, "token": tok}; c.update({"lb-instance-id": lb} if lb else {}); return c
    def _h(tok, lb): h = {"Authorization": tok, "Content-Type": "application/x-www-form-urlencoded", "Origin": BASE_URL, "Referer": BASE_URL}; h.update({"lb-instance-id": lb} if lb else {}); return h
    cookies, headers = _c(token, jsessionid, lb), _h(token, lb)
    for attempt in range(retries):
        try:
            resp = requests.post(DATA_URL, data=payload, headers=headers, cookies=cookies, timeout=90)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.HTTPError:
            if resp.status_code == 403 and attempt < retries - 1:
                print(f"    ⚠ 403 — re-logging in..."); time.sleep(10)
                fresh = _aera_login(); token = fresh["access_token"]
                jsessionid = fresh.get("jsessionid", jsessionid); lb = fresh.get("lb_instance_id", lb)
                cookies, headers = _c(token, jsessionid, lb), _h(token, lb); continue
            raise
        except requests.exceptions.ConnectionError:
            if attempt < retries - 1: time.sleep(10); continue
            raise


# ── Fetch one Lag-1 pair ──────────────────────────────────────────────────────

def fetch_lag_pair(snapshot_month, forecast_month, token, jsessionid, lb):
    """Returns df with Country_Name, Material_Number, Adjusted_FC aggregated."""
    print(f"  Fetching Lag-1 {forecast_month} (snapshot={snapshot_month})...", end=" ", flush=True)
    t0 = time.time()

    base = {
        "sheetid": SHEETID, "bioid": BIOID, "fid": FID,
        "row": ROW, "col": "", "mea": MEA,
        "filter": f"{SNAPSHOT_DIM}~cnt|{snapshot_month}~EN^{FORECAST_DIM}~cnt|{forecast_month}~EN",
        "pot": "0", "sort": FORECAST_DIM, "dir": "ASC",
        "in_val": "[]", "source": "report",
        "uom": "", "currency": "", "rate": "", "currencyDate": "T",
        "plimit": str(PAGE_LIMIT),
    }

    raw = _post(token, jsessionid, {**base, "pstart": "0", "page": "1"}, lb)
    total = raw.get("totalRows", 0)
    if total == 0:
        print(f"0 rows")
        return pd.DataFrame(columns=["Country_Name", "Material_Number", "Adjusted_FC"])

    pages = max(1, math.ceil(total / PAGE_LIMIT))
    all_rows = list(raw["data"])
    for p in range(1, pages):
        r = _post(token, jsessionid, {**base, "pstart": str(p * PAGE_LIMIT), "page": str(p + 1)}, lb)
        all_rows.extend(r["data"])

    col_ids = [f["name"] for f in raw["metaData"]["fields"]]
    df = pd.DataFrame(all_rows, columns=col_ids)

    # Rename to friendly names
    country_id   = "060CACC0-EC7A-4946-A601-C86E4D69AB29_DAB42D7F-6407-4CA3-ADA4-92456C940A47"
    customer_id  = "50191AB8-1EDB-11ED-A548-0A617A24E20D_9EB3B832-5F1A-4FF1-814B-CC82933E9F14"
    material_id  = "04E2EDB1-1A36-11ED-A548-0A617A24E20D_454245B2-6AF3-49B8-AA8E-18FEC4E340DC"
    adjfc_id     = "E3524740-47C7-4C30-A381-333FC13DEBD6|SUM|"
    df = df.rename(columns={
        country_id: "Country_Name", customer_id: "Customer_Number",
        material_id: "Material_Number", adjfc_id: "Adjusted_FC"
    })
    df["Adjusted_FC"] = pd.to_numeric(df["Adjusted_FC"], errors="coerce").fillna(0)

    # Group by Material + Country + Customer
    agg = df.groupby(["Material_Number", "Country_Name", "Customer_Number"], as_index=False)["Adjusted_FC"].sum()
    agg = agg[agg["Adjusted_FC"] != 0]

    print(f"{total:,} rows → {len(agg):,} mat×country combos ({time.time()-t0:.0f}s)")
    return agg


# ── Actuals from adjfc parquet ────────────────────────────────────────────────

def load_actuals():
    path = os.path.join(DIR, "adjfc_nz.parquet")
    df = pd.read_parquet(path)

    # Use friendly column names (already mapped in fetch_adjfc.py)
    act = df[df["Month Year"].isin(FORECAST_MONTHS)][
        ["Material Number", "Country Name", "Customer Number", "Month Year", "Actuals"]
    ].copy()
    act.columns = ["Material_Number", "Country_Name", "Customer_Number", "Month_Year", "Actuals"]
    act["Actuals"] = pd.to_numeric(act["Actuals"], errors="coerce").fillna(0)
    act["Customer_Number"] = act["Customer_Number"].astype(str)

    agg = act.groupby(["Material_Number", "Country_Name", "Customer_Number", "Month_Year"], as_index=False)["Actuals"].sum()
    agg = agg[agg["Actuals"] != 0]

    # Pivot wide
    pivot = agg.pivot_table(
        index=["Material_Number", "Country_Name", "Customer_Number"],
        columns="Month_Year",
        values="Actuals",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    # Ensure all 5 months exist
    for m in FORECAST_MONTHS:
        if m not in pivot.columns:
            pivot[m] = 0.0

    # Rename to Actual_*
    rename = {m: f"Actual_{m.replace(' ', '_')}" for m in FORECAST_MONTHS}
    pivot = pivot.rename(columns=rename)
    return pivot


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    t_start = time.time()
    tok = ensure_token(min_seconds=300)
    token, jsessionid, lb = tok["access_token"], tok.get("jsessionid", ""), tok.get("lb_instance_id", "")

    KEY = ["Material_Number", "Country_Name", "Customer_Number"]

    # ── Fetch all 5 Lag-1 pairs ───────────────────────────────────────────────
    print("Fetching Lag-1 forecasts (5 months)...")
    lag1_frames = []
    for snapshot, forecast in LAG1_PAIRS:
        df = fetch_lag_pair(snapshot, forecast, token, jsessionid, lb)
        df = df.rename(columns={"Adjusted_FC": f"Lag1_{forecast.replace(' ', '_')}"})
        lag1_frames.append(df)

    # ── Fetch all 5 Lag-3 pairs ───────────────────────────────────────────────
    print("\nFetching Lag-3 forecasts (5 months)...")
    lag3_frames = []
    for snapshot, forecast in LAG3_PAIRS:
        df = fetch_lag_pair(snapshot, forecast, token, jsessionid, lb)
        df = df.rename(columns={"Adjusted_FC": f"Lag3_{forecast.replace(' ', '_')}"})
        lag3_frames.append(df)

    # ── Merge all lag frames ──────────────────────────────────────────────────
    print("\nMerging lag frames...")
    lag_wide = lag1_frames[0]
    for frame in lag1_frames[1:] + lag3_frames:
        lag_wide = lag_wide.merge(frame, on=KEY, how="outer")

    lag1_col_names = [f"Lag1_{m.replace(' ', '_')}" for m in FORECAST_MONTHS]
    lag3_col_names = [f"Lag3_{m.replace(' ', '_')}" for m in FORECAST_MONTHS]
    lag_col_names  = lag1_col_names + lag3_col_names
    for c in lag_col_names:
        lag_wide[c] = lag_wide[c].fillna(0)

    # ── Load actuals ──────────────────────────────────────────────────────────
    print("Loading actuals from adjfc parquet...")
    actuals = load_actuals()
    actual_col_names = [f"Actual_{m.replace(' ', '_')}" for m in FORECAST_MONTHS]

    # ── Join ──────────────────────────────────────────────────────────────────
    # Ensure Customer_Number is string in both before joining
    lag_wide["Customer_Number"] = lag_wide["Customer_Number"].astype(str)
    actuals["Customer_Number"] = actuals["Customer_Number"].astype(str)
    result = lag_wide.merge(actuals, on=KEY, how="outer")
    for c in lag_col_names + actual_col_names:
        result[c] = result[c].fillna(0)

    # Keep only rows with any non-zero value
    value_cols = lag_col_names + actual_col_names
    result = result[(result[value_cols] != 0).any(axis=1)]

    # Sort
    result = result.sort_values(["Country_Name", "Customer_Number", "Material_Number"]).reset_index(drop=True)

    col_order = ["Material_Number", "Country_Name", "Customer_Number"] + lag1_col_names + lag3_col_names + actual_col_names
    result = result[col_order]

    # ── Save ──────────────────────────────────────────────────────────────────
    out_xlsx = os.path.join(DIR, "lag1_vs_actuals.xlsx")
    out_parquet = os.path.join(DIR, "lag1_vs_actuals.parquet")

    result.to_parquet(out_parquet, index=False)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="Lag1 vs Actuals")
        ws = writer.sheets["Lag1 vs Actuals"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Header formatting
        lag_fill    = PatternFill("solid", fgColor="C6EFCE")   # light green — lag
        actual_fill = PatternFill("solid", fgColor="BDD7EE")   # light blue — actual
        dim_fill    = PatternFill("solid", fgColor="D9D9D9")   # grey — dimension
        bold        = Font(bold=True)
        center      = Alignment(horizontal="center")
        thin        = Side(style="thin")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(result.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = bold
            cell.alignment = center
            cell.border = border
            if col_name.startswith("Lag1_"):
                cell.fill = lag_fill
            elif col_name.startswith("Actual_"):
                cell.fill = actual_fill
            else:
                cell.fill = dim_fill

        # Number format for value columns
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                col_name = result.columns[cell.column - 1]
                if col_name.startswith(("Lag1_", "Actual_")):
                    cell.number_format = "#,##0.0"
                cell.border = border

        # Column widths
        ws.column_dimensions["A"].width = 18  # Material_Number
        ws.column_dimensions["B"].width = 20  # Country_Name
        ws.column_dimensions["C"].width = 18  # Customer_Number
        for i in range(4, len(result.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 16

        ws.freeze_panes = "C2"

    elapsed = (time.time() - t_start) / 60
    print(f"\n✓ Done in {elapsed:.1f} min")
    print(f"  Rows: {len(result):,}  |  Mat×Country combos")
    print(f"  Saved → {out_xlsx}")
    print(f"\nSummary totals:")
    for c in lag1_col_names + lag3_col_names + actual_col_names:
        print(f"  {c}: {result[c].sum():>12,.1f}")


if __name__ == "__main__":
    main()
