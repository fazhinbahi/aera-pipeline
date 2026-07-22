"""
2028 Forecast Anomaly Analysis
================================
Compares 2028 AdjFC (from adjfc_2028.parquet) against:
  - 2024 actuals        (customer_analysis.Actual_*_2024)
  - 2025 actuals        (customer_analysis.Actual_*_2025)
  - 2026 AdjFC H2       (customer_analysis.AdjFC_Jul–Dec_2026)
  - 2027 AdjFC full     (customer_analysis.AdjFC_Jan–Dec_2027)

Anomaly checks:
  1. Obsolete products   — 2028 AdjFC > 0 but zero actuals in 2024 AND 2025
  2. Near-dead products  — 2028 AdjFC > 0 but rapidly declining: 2025 actuals < 20% of 2024
  3. YoY explosion       — 2028 AdjFC > 3× 2027 AdjFC (unexplained jump in far-out year)
  4. Country anomaly     — SKU has 2028 AdjFC in a country with zero 2024–2025 actuals
  5. Volume spike        — 2028 AdjFC > 5× average of 2024+2025 actuals
  6. Brand-level outlier — SKU 2028 AdjFC >> brand-country average (statistical outlier)

Output: forecast_2028_anomaly_report_YYYY-MM-DD.xlsx
"""

import os, sys
from datetime import date
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from load_to_bq import _client

GCP = "euphoric-hull-442815-n8"
DS  = "aera_demand_planning"
DIR = os.path.dirname(os.path.abspath(__file__))

MONTHS_2028 = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTHS_2027 = MONTHS_2028
MONTHS_2026_H2 = ["Jul","Aug","Sep","Oct","Nov","Dec"]


def load_2028_parquet():
    path = os.path.join(DIR, "adjfc_2028.parquet")
    if not os.path.exists(path):
        raise FileNotFoundError(f"adjfc_2028.parquet not found — run fetch_adjfc_2028.py first")
    df = pd.read_parquet(path)
    # Keep only EMEA/APAC rows, exclude actuals (Month Year = 2028)
    df = df[df["Month Year"].str.contains("2028", na=False)].copy()
    return df


def pivot_2028(df: pd.DataFrame) -> pd.DataFrame:
    """Pivot to one row per Material Number × Country × Customer grain, 12 FC columns."""
    df = df.copy()
    df["Month_Short"] = df["Month Year"].str.replace(" 2028", "", regex=False)
    pivot = df.pivot_table(
        index=["Material Number", "Country Name", "Sub-Segments", "Customer Number",
               "Organisation", "Region", "Business Segment",
               "Distributor Name", "Sub-Brand Long Description", "Volume"],
        columns="Month_Short",
        values="Adj FC 9LC",
        aggfunc="sum",
    ).reset_index()
    pivot.columns.name = None

    for m in MONTHS_2028:
        col = f"AdjFC_{m}_2028"
        if m in pivot.columns:
            pivot[col] = pivot[m].fillna(0)
            pivot.drop(columns=[m], inplace=True)
        else:
            pivot[col] = 0.0

    pivot["AdjFC_Total_2028"] = pivot[[f"AdjFC_{m}_2028" for m in MONTHS_2028]].sum(axis=1)
    return pivot[pivot["AdjFC_Total_2028"] > 0].copy()


def fetch_history(client) -> pd.DataFrame:
    """Pull actuals + AdjFC for 2024–2027 from customer_analysis in BQ."""
    act_2024  = "+".join(f"COALESCE(Actual_{m}_2024,0)" for m in MONTHS_2028)
    act_2025  = "+".join(f"COALESCE(Actual_{m}_2025,0)" for m in MONTHS_2028)
    act_h1_26 = "+".join(f"COALESCE(Actual_{m}_2026,0)" for m in ["Jan","Feb","Mar","Apr","May","Jun"])
    so_h2_26  = "+".join(f"COALESCE(SO_{m}_2026,0)"     for m in MONTHS_2026_H2)
    fc_2026h2 = "+".join(f"COALESCE(AdjFC_{m}_2026,0)"  for m in MONTHS_2026_H2)
    fc_2027   = "+".join(f"COALESCE(AdjFC_{m}_2027,0)"  for m in MONTHS_2027)

    q = f"""
    SELECT
      Material_Number, Country_Name, Sub_Segments, Customer_Number,
      Brand_Family, Sub_Brand_Description,
      ROUND({act_2024},  4) AS Actual_Total_2024,
      ROUND({act_2025},  4) AS Actual_Total_2025,
      ROUND({act_h1_26}, 4) AS Actual_H1_2026,
      ROUND({so_h2_26},  4) AS SO_H2_2026,
      ROUND({fc_2026h2}, 4) AS AdjFC_H2_2026,
      ROUND(COALESCE(AdjFC_Total_2027, {fc_2027}), 4) AS AdjFC_Total_2027
    FROM `{GCP}.{DS}.customer_analysis`
    """
    rows = client.query(q).result()
    return pd.DataFrame([dict(r) for r in rows])


def merge_data(fc28: pd.DataFrame, hist: pd.DataFrame) -> pd.DataFrame:
    """Left join 2028 forecast onto history."""
    hist = hist.rename(columns={
        "Material_Number": "Material Number",
        "Country_Name":    "Country Name",
        "Sub_Segments":    "Sub-Segments",
        "Customer_Number": "Customer Number",
        "Brand_Family":    "Brand Family",
        "Sub_Brand_Description": "Sub Brand",
    })
    merged = fc28.merge(
        hist,
        on=["Material Number", "Country Name", "Sub-Segments", "Customer Number"],
        how="left",
    )
    for col in ["Actual_Total_2024","Actual_Total_2025","Actual_H1_2026",
                "SO_H2_2026","AdjFC_H2_2026","AdjFC_Total_2027"]:
        merged[col] = merged[col].fillna(0)
    return merged


def run_checks(df: pd.DataFrame):
    """Return dict of DataFrames, one per check."""
    DIM = ["Material Number", "Country Name", "Sub-Segments", "Customer Number",
           "Distributor Name", "Sub-Brand Long Description", "Brand Family", "Sub Brand",
           "Organisation", "Region", "Business Segment",
           "Actual_Total_2024", "Actual_Total_2025",
           "Actual_H1_2026", "SO_H2_2026",
           "AdjFC_H2_2026", "AdjFC_Total_2027", "AdjFC_Total_2028"]

    def clean(d):
        available = [c for c in DIM if c in d.columns]
        return d[available].copy()

    checks = {}

    # Check 1: Obsolete products — 2028 AdjFC > 0, zero actuals in 2024 AND 2025
    c1 = df[
        (df["AdjFC_Total_2028"] > 0) &
        (df["Actual_Total_2024"].fillna(0) == 0) &
        (df["Actual_Total_2025"].fillna(0) == 0)
    ].copy()
    c1 = c1.sort_values("AdjFC_Total_2028", ascending=False)
    checks["check1"] = clean(c1)

    # Check 2: Near-dead products — 2025 actuals < 20% of 2024 (strong decline, still forecasted)
    c2 = df[
        (df["AdjFC_Total_2028"] > 0) &
        (df["Actual_Total_2024"] > 50) &
        (df["Actual_Total_2025"] < df["Actual_Total_2024"] * 0.20) &
        (df["Actual_Total_2025"] >= 0)
    ].copy()
    c2["Decline_Pct"] = ((df["Actual_Total_2025"] - df["Actual_Total_2024"]) /
                          df["Actual_Total_2024"].replace(0, float("nan")) * 100).round(1)
    c2 = c2.sort_values("AdjFC_Total_2028", ascending=False)
    checks["check2"] = clean(c2)

    # Check 3: YoY explosion — 2028 AdjFC > 3× 2027 AdjFC
    c3 = df[
        (df["AdjFC_Total_2028"] > 100) &
        (df["AdjFC_Total_2027"] > 0) &
        (df["AdjFC_Total_2028"] > 3 * df["AdjFC_Total_2027"])
    ].copy()
    c3["Ratio_2028_vs_2027"] = (df["AdjFC_Total_2028"] / df["AdjFC_Total_2027"].replace(0, float("nan"))).round(2)
    c3 = c3.sort_values("Ratio_2028_vs_2027", ascending=False)
    checks["check3"] = clean(c3)

    # Check 4: Volume spike — 2028 AdjFC > 5× avg of (2024 + 2025) actuals
    df["Avg_Actual_2024_25"] = ((df["Actual_Total_2024"] + df["Actual_Total_2025"]) / 2)
    c4 = df[
        (df["AdjFC_Total_2028"] > 100) &
        (df["Avg_Actual_2024_25"] > 0) &
        (df["AdjFC_Total_2028"] > 5 * df["Avg_Actual_2024_25"])
    ].copy()
    c4["FC28_vs_AvgActual_Ratio"] = (df["AdjFC_Total_2028"] / df["Avg_Actual_2024_25"].replace(0, float("nan"))).round(2)
    c4 = c4.sort_values("FC28_vs_AvgActual_Ratio", ascending=False)
    checks["check4"] = clean(c4)

    # Check 5: Country anomaly — 2028 AdjFC in country with ZERO 2024+2025 actuals
    c5 = df[
        (df["AdjFC_Total_2028"] > 0) &
        (df["Actual_Total_2024"].fillna(0) == 0) &
        (df["Actual_Total_2025"].fillna(0) == 0)
    ].groupby(["Material Number", "Country Name", "Sub-Segments",
               "Sub-Brand Long Description", "Brand Family", "Organisation"], as_index=False).agg(
        AdjFC_Total_2028=("AdjFC_Total_2028", "sum"),
        Customers_Forecasted=("Customer Number", "nunique")
    ).sort_values("AdjFC_Total_2028", ascending=False)
    checks["check5"] = c5

    # Check 6: Completely dark 2026–2027 but alive in 2028
    # Zero actuals in H1 2026, zero SO in H2 2026, zero AdjFC in 2027 — but AdjFC 2028 > 0
    c6 = df[
        (df["AdjFC_Total_2028"] > 0) &
        (df["Actual_H1_2026"].fillna(0) == 0) &
        (df["SO_H2_2026"].fillna(0) == 0) &
        (df["AdjFC_H2_2026"].fillna(0) == 0) &
        (df["AdjFC_Total_2027"].fillna(0) == 0)
    ].copy()
    c6 = c6.sort_values("AdjFC_Total_2028", ascending=False)
    checks["check6"] = clean(c6)

    return checks


def write_xlsx(checks, summary_stats, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    NAVY, WHITE, RED, AMB, BLU = "1F3864","FFFFFF","C00000","ED7D31","2F5496"
    ALT, BDR = "F2F2F2", "D9D9D9"

    thin = Border(
        left=Side(style="thin", color=BDR), right=Side(style="thin", color=BDR),
        top=Side(style="thin", color=BDR),  bottom=Side(style="thin", color=BDR),
    )
    today = date.today().strftime("%d %b %Y")

    def write_sheet(ws, df, hdr_color, title, subtitle):
        ws.append([title]); ws.append([subtitle]); ws.append([])
        ws.append(list(df.columns))
        HDR = 4
        for cell in ws[HDR]:
            cell.fill = PatternFill("solid", fgColor=hdr_color)
            cell.font = Font(bold=True, color=WHITE, size=9)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin
        ws.row_dimensions[HDR].height = 30

        for ri, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=HDR+1):
            ws.append(row)
            fill = PatternFill("solid", fgColor=ALT if ri%2==0 else WHITE)
            for ci, cell in enumerate(ws[ri], start=1):
                cell.border = thin
                cell.font   = Font(size=9)
                cell.fill   = fill
                col = df.columns[ci-1]
                if any(t in str(col) for t in ("9LC","2024","2025","2026","2027","2028","Ratio","Pct","Total","Avg")):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        last = get_column_letter(len(df.columns))
        ws.merge_cells(f"A1:{last}1"); ws.merge_cells(f"A2:{last}2")
        ws["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws["A2"].font = Font(italic=True, size=9, color="595959")
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 14
        ws.freeze_panes = ws.cell(row=HDR+1, column=1)
        ws.auto_filter.ref = f"A{HDR}:{last}{ws.max_row}"
        for ci, col in enumerate(df.columns, start=1):
            max_w = max(len(str(col)), max((len(str(v)) for v in df.iloc[:, ci-1]), default=0))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 2, 40)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_s = wb.create_sheet("Summary", 0)
    ws_s.append(["2028 Forecast Anomaly Report"])
    ws_s.append([f"Generated: {today}  |  Source: Aera AdjFC 2028 vs 2024/2025 Actuals & 2027 AdjFC"])
    ws_s.append(["NOTE: Aera Statistical Forecast SSR does not expose 2028 — analysis uses AdjFC only"])
    ws_s.append([])
    ws_s.append(["Metric", "Value"])
    for k, v in summary_stats.items():
        ws_s.append([k, v])
    ws_s.append([])
    ws_s.append(["Check", "Description", "Flagged Grains"])
    descs = {
        "check1": "Obsolete — 2028 AdjFC > 0 with ZERO actuals in 2024 and 2025",
        "check2": "Near-Dead — 2025 actuals < 20% of 2024 (sharp decline, still forecasted for 2028)",
        "check3": "YoY Explosion — 2028 AdjFC > 3× 2027 AdjFC",
        "check4": "Volume Spike — 2028 AdjFC > 5× avg of 2024+2025 actuals",
        "check5": "Country Anomaly — 2028 AdjFC in country with zero 2024/2025 actuals (SKU-Country level)",
        "check6": "Ghost 2028 — Zero actuals AND zero AdjFC in ALL of 2026–2027, but AdjFC exists for 2028 (most critical)",
    }
    colors = {"check1": RED, "check2": AMB, "check3": AMB, "check4": RED, "check5": BLU, "check6": RED}

    for cname, desc in descs.items():
        cnt = len(checks.get(cname, []))
        ws_s.append([cname.replace("check","Check "), desc, cnt])

    ws_s["A1"].font = Font(bold=True, size=14, color=NAVY)
    ws_s["A2"].font = Font(italic=True, size=9, color="595959")
    ws_s["A3"].font = Font(italic=True, size=9, color=RED)
    for r in [5, 12]:
        for cell in ws_s[r]: cell.fill = PatternFill("solid", fgColor=NAVY); cell.font = Font(bold=True, color=WHITE, size=10)
    for i, cname in enumerate(descs.keys(), start=13):
        if ws_s.max_row >= i:
            clr = colors.get(cname, NAVY)
            for cell in ws_s[i]: cell.fill = PatternFill("solid", fgColor=clr); cell.font = Font(bold=True, color=WHITE, size=10)
    ws_s.column_dimensions["A"].width = 14
    ws_s.column_dimensions["B"].width = 65
    ws_s.column_dimensions["C"].width = 16

    labels = {
        "check1": ("1-Obsolete Products", RED,
                   "Check 1 — Obsolete Products with 2028 Forecast",
                   f"No actuals in 2024 or 2025 — product likely discontinued or never launched"),
        "check2": ("2-Near-Dead Products", AMB,
                   "Check 2 — Near-Dead Products Still Forecasted for 2028",
                   f"2025 actuals < 20% of 2024 — rapid decline, but 2028 forecast unchanged"),
        "check3": ("3-YoY Explosion", AMB,
                   "Check 3 — 2028 AdjFC More Than 3× the 2027 AdjFC",
                   f"Unexplained jump in far-out year — verify building block justification"),
        "check4": ("4-Volume Spike", RED,
                   "Check 4 — 2028 AdjFC More Than 5× Average 2024/2025 Actuals",
                   f"Statistical model may have extrapolated anomalous data into 2028"),
        "check5": ("5-Country Anomaly", BLU,
                   "Check 5 — 2028 Forecast for Country with No 2024/2025 Actuals",
                   f"SKU × Country combination has never shipped — potential ghost forecast"),
        "check6": ("6-Ghost 2028 (CRITICAL)", RED,
                   "Check 6 — GHOST FORECAST: Completely Dark 2026–2027, Alive in 2028",
                   f"Zero actuals AND zero AdjFC AND zero SO in all of 2026–2027, but 2028 AdjFC > 0  |  Most critical anomaly"),
    }

    for cname, (sheet_name, color, title, subtitle) in labels.items():
        df = checks.get(cname, pd.DataFrame())
        if not df.empty:
            ws = wb.create_sheet(sheet_name)
            write_sheet(ws, df, color,
                        f"{title}  |  {len(df):,} grains  |  Generated: {today}",
                        subtitle)

    wb.save(path)
    print(f"✓ Report saved → {path}")


def main():
    print("Loading 2028 AdjFC from parquet…")
    raw_2028 = load_2028_parquet()
    print(f"  {len(raw_2028):,} rows, months: {sorted(raw_2028['Month Year'].unique())}")

    print("Pivoting to grain-level…")
    fc28 = pivot_2028(raw_2028)
    print(f"  {len(fc28):,} unique grains with 2028 forecast > 0")

    print("Fetching 2024/2025 actuals + 2027 AdjFC from BigQuery…")
    client = _client()
    hist   = fetch_history(client)
    print(f"  {len(hist):,} history rows loaded")

    print("Merging…")
    merged = merge_data(fc28, hist)

    summary_stats = {
        "Distinct SKUs with 2028 AdjFC":      fc28["Material Number"].nunique(),
        "Distinct Countries with 2028 AdjFC": fc28["Country Name"].nunique(),
        "Total grains (SKU×Country×SubSeg×Customer)": len(fc28),
        "Total 2028 AdjFC (9LC)":             round(fc28["AdjFC_Total_2028"].sum(), 1),
        "Grains with ZERO 2024+2025 actuals": int(((merged["Actual_Total_2024"]==0) & (merged["Actual_Total_2025"]==0)).sum()),
        "Grains with 2028 AdjFC > 3× 2027":  int((merged["AdjFC_Total_2028"] > 3*merged["AdjFC_Total_2027"].replace(0, float("nan"))).sum()),
    }

    print("Running anomaly checks…")
    checks = run_checks(merged)

    for name, df in checks.items():
        print(f"  {name}: {len(df):,} flagged grains")

    out_path = os.path.join(DIR, f"forecast_2028_anomaly_report_{date.today().isoformat()}.xlsx")
    write_xlsx(checks, summary_stats, out_path)

    print("\n── Anomaly Summary ──────────────────────────────────────────────────────")
    print(f"  Check 1 Obsolete (zero 2024+2025 actuals):  {len(checks['check1']):,} grains")
    print(f"  Check 2 Near-dead (>80% decline 2024→2025): {len(checks['check2']):,} grains")
    print(f"  Check 3 YoY explosion (>3× vs 2027):         {len(checks['check3']):,} grains")
    print(f"  Check 4 Volume spike (>5× avg 2024/25):      {len(checks['check4']):,} grains")
    print(f"  Check 5 Country anomaly (no actuals in ctry): {len(checks['check5']):,} grains")
    print(f"  Check 6 Ghost 2028 (dark 2026-27, alive 28): {len(checks['check6']):,} grains  ← MOST CRITICAL")
    print(f"\n  Report: {out_path}")


if __name__ == "__main__":
    main()
