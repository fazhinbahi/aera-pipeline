"""
SO vs Forecast Mismatch Report
===============================
Grain: SKU × Country × Customer  (one row per grain)
Months (Jul–Dec 2026) become column headers — SO, FC, Gap per month.

Layer 1 (priority): rows where at least one month has Sales Order > AdjFC
Layer 2:            rows where at least one month has AdjFC > SO  (SO > 0)

Output: so_fc_mismatch_report_YYYY-MM-DD.xlsx
"""

import os
import sys
from datetime import date

import pandas as pd
from google.cloud import bigquery

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from load_to_bq import _client

GCP_PROJECT = "euphoric-hull-442815-n8"
DATASET     = "aera_demand_planning"
OPEN_MONTHS = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

DIM_COLS = [
    "Business Segment", "Sub Segment", "Country",
    "Customer Number", "Customer Name",
    "Material Number", "Sub Brand", "Brand Family",
]


def _fetch_data(client: bigquery.Client) -> pd.DataFrame:
    """Pull all grains that have at least one mismatch month — wide format."""
    ca = f"`{GCP_PROJECT}.{DATASET}.customer_analysis`"
    fc = f"`{GCP_PROJECT}.{DATASET}.adjfc_raw`"

    # WHERE: any month has SO≠FC mismatch (SO>FC  OR  FC>SO with SO>0)
    mismatch_clauses = []
    for m in OPEN_MONTHS:
        mismatch_clauses.append(
            f"(ABS(SO_{m}_2026 - AdjFC_{m}_2026) > 0.0001 "
            f"AND (SO_{m}_2026 > AdjFC_{m}_2026 OR (AdjFC_{m}_2026 > SO_{m}_2026 AND SO_{m}_2026 > 0)))"
        )
    where = " OR ".join(mismatch_clauses)

    # Select dimension + all 6 pairs of SO/FC columns
    month_cols = ", ".join(
        f"ca.SO_{m}_2026, ca.AdjFC_{m}_2026" for m in OPEN_MONTHS
    )

    q = f"""
    WITH base AS (
        SELECT
            ca.Business_Segment,
            ca.Sub_Segments,
            ca.Country_Name,
            ca.Customer_Number,
            ca.Material_Number,
            ca.Sub_Brand_Description,
            ca.Brand_Family,
            {month_cols}
        FROM {ca} ca
        WHERE {where}
    ),
    customer_names AS (
        SELECT DISTINCT Customer_Number,
               COALESCE(NULLIF(TRIM(Distributor_Name), ''), Customer_Number) AS Distributor_Name
        FROM {fc}
        WHERE Customer_Number IS NOT NULL AND Customer_Number != ''
    )
    SELECT
        b.Business_Segment,
        b.Sub_Segments,
        b.Country_Name,
        b.Customer_Number,
        COALESCE(cn.Distributor_Name, b.Customer_Number) AS Customer_Name,
        b.Material_Number,
        b.Sub_Brand_Description,
        b.Brand_Family,
        {", ".join(f"b.SO_{m}_2026, b.AdjFC_{m}_2026" for m in OPEN_MONTHS)}
    FROM base b
    LEFT JOIN customer_names cn ON b.Customer_Number = cn.Customer_Number
    ORDER BY b.Sub_Segments, b.Country_Name, b.Customer_Number, b.Material_Number
    """

    return client.query(q).to_dataframe()


def _build_wide(raw: pd.DataFrame):
    """Rename columns, compute gap per month, split into Layer 1 / Layer 2."""
    df = raw.rename(columns={
        "Business_Segment":      "Business Segment",
        "Sub_Segments":          "Sub Segment",
        "Country_Name":          "Country",
        "Customer_Number":       "Customer Number",
        "Customer_Name":         "Customer Name",
        "Material_Number":       "Material Number",
        "Sub_Brand_Description": "Sub Brand",
        "Brand_Family":          "Brand Family",
    })

    # Rename SO/FC columns to readable month labels and add Gap columns
    ordered_month_cols = []
    for m in OPEN_MONTHS:
        so_raw  = f"SO_{m}_2026"
        fc_raw  = f"AdjFC_{m}_2026"
        so_lbl  = f"{m} 2026 SO"
        fc_lbl  = f"{m} 2026 FC"
        gap_lbl = f"{m} 2026 Gap"
        df = df.rename(columns={so_raw: so_lbl, fc_raw: fc_lbl})
        df[gap_lbl] = (df[so_lbl] - df[fc_lbl]).round(4)
        ordered_month_cols += [so_lbl, fc_lbl, gap_lbl]

    final_cols = DIM_COLS + ordered_month_cols

    # Layer 1: at least one month where SO > FC
    l1_mask = pd.Series(False, index=df.index)
    for m in OPEN_MONTHS:
        l1_mask |= df[f"{m} 2026 SO"] > df[f"{m} 2026 FC"]

    # Layer 2: at least one month where FC > SO  (and that month's SO > 0)
    l2_mask = pd.Series(False, index=df.index)
    for m in OPEN_MONTHS:
        l2_mask |= (df[f"{m} 2026 FC"] > df[f"{m} 2026 SO"]) & (df[f"{m} 2026 SO"] > 0)

    layer1 = df[l1_mask][final_cols].copy()
    layer2 = df[l2_mask][final_cols].copy()

    # Sort Layer 1 by total gap descending (most urgent first)
    gap_cols = [f"{m} 2026 Gap" for m in OPEN_MONTHS]
    layer1["_total_gap"] = layer1[gap_cols].clip(lower=0).sum(axis=1)
    layer1 = layer1.sort_values("_total_gap", ascending=False).drop(columns="_total_gap")

    layer2["_total_neg"] = layer2[gap_cols].clip(upper=0).sum(axis=1)
    layer2 = layer2.sort_values("_total_neg", ascending=True).drop(columns="_total_neg")

    return layer1, layer2


def _write_xlsx(layer1: pd.DataFrame, layer2: pd.DataFrame, path: str):
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY       = "1F3864"
    WHITE      = "FFFFFF"
    L1_HDR     = "C00000"   # deep red  — Layer 1 header
    L1_ALT     = "FDECEA"   # blush     — Layer 1 alternating
    L2_HDR     = "ED7D31"   # amber     — Layer 2 header
    L2_ALT     = "FFF2CC"   # pale gold — Layer 2 alternating
    FLAG_RED   = "FF0000"   # cell-level flag: SO > FC
    FLAG_AMB   = "FFC000"   # cell-level flag: FC > SO
    BORDER_C   = "D9D9D9"

    thin = Border(
        left=Side(style="thin", color=BORDER_C),
        right=Side(style="thin", color=BORDER_C),
        top=Side(style="thin", color=BORDER_C),
        bottom=Side(style="thin", color=BORDER_C),
    )

    # Map month Gap column index → (SO col idx, FC col idx) for cell-level highlighting
    n_dim = len(DIM_COLS)  # 8 dim columns

    def _month_gap_col_positions():
        """Return {gap_col_index_1based: (so_1based, fc_1based)} for all months."""
        pos = {}
        for i, m in enumerate(OPEN_MONTHS):
            base = n_dim + i * 3 + 1        # 1-based column of SO for this month
            so_i, fc_i, gap_i = base, base + 1, base + 2
            pos[gap_i] = (so_i, fc_i)
        return pos

    gap_positions = _month_gap_col_positions()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _write_sheet(ws, df: pd.DataFrame, hdr_color: str, alt_color: str,
                     title: str, subtitle: str):
        n_cols = len(df.columns)

        # Row 1: title, Row 2: subtitle, Row 3: blank
        ws.append([title])
        ws.append([subtitle])
        ws.append([])
        HDR_ROW = 4

        # ── Header row ────────────────────────────────────────────────────────
        ws.append(list(df.columns))
        for cell in ws[HDR_ROW]:
            cell.fill      = PatternFill("solid", fgColor=hdr_color)
            cell.font      = Font(bold=True, color=WHITE, size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = thin

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_idx, row_vals in enumerate(
            dataframe_to_rows(df, index=False, header=False), start=HDR_ROW + 1
        ):
            ws.append(row_vals)
            base_fill = PatternFill("solid", fgColor=alt_color if row_idx % 2 == 0 else WHITE)

            for col_idx, cell in enumerate(ws[row_idx], start=1):
                col_name = df.columns[col_idx - 1]
                cell.border = thin
                cell.font   = Font(size=9)

                # Gap columns: colour by sign
                if col_idx in gap_positions:
                    val = cell.value
                    if val is not None and val > 0.0001:
                        cell.fill = PatternFill("solid", fgColor="FFDEDE")  # light red
                        cell.font = Font(size=9, bold=True, color=FLAG_RED)
                    elif val is not None and val < -0.0001:
                        cell.fill = PatternFill("solid", fgColor="FFF3CD")  # light amber
                        cell.font = Font(size=9, bold=True, color="7F5000")
                    else:
                        cell.fill = base_fill
                else:
                    cell.fill = base_fill

                # Number formats
                if any(tag in col_name for tag in (" SO", " FC", " Gap")):
                    cell.number_format = "#,##0.00"
                    cell.alignment     = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")

        # ── Title styling ─────────────────────────────────────────────────────
        last_col_ltr = get_column_letter(n_cols)
        ws.merge_cells(f"A1:{last_col_ltr}1")
        ws.merge_cells(f"A2:{last_col_ltr}2")
        ws["A1"].font = Font(bold=True, size=13, color=NAVY)
        ws["A2"].font = Font(italic=True, size=9, color="595959")
        ws.row_dimensions[1].height   = 22
        ws.row_dimensions[2].height   = 14
        ws.row_dimensions[HDR_ROW].height = 32

        # ── Column widths ─────────────────────────────────────────────────────
        dim_widths = [16, 18, 18, 16, 28, 16, 26, 18]
        for i, w in enumerate(dim_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Month columns: SO=11, FC=11, Gap=9
        for i, m in enumerate(OPEN_MONTHS):
            base = n_dim + i * 3 + 1
            ws.column_dimensions[get_column_letter(base)].width     = 11
            ws.column_dimensions[get_column_letter(base + 1)].width = 11
            ws.column_dimensions[get_column_letter(base + 2)].width = 9

        # ── Freeze & filter ───────────────────────────────────────────────────
        ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)
        ws.auto_filter.ref = f"A{HDR_ROW}:{last_col_ltr}{ws.max_row}"

    today_str = date.today().strftime("%d %b %Y")

    _write_sheet(
        wb.create_sheet("Layer 1 — SO > Forecast"), layer1,
        hdr_color=L1_HDR, alt_color=L1_ALT,
        title="Layer 1 — Sales Order Exceeds Forecast  [PRIORITY FLAGS]",
        subtitle=f"Grain: SKU × Country × Customer  |  Open months: Jul–Dec 2026  |  Generated: {today_str}  |  {len(layer1):,} records  |  Red gap = SO > FC",
    )

    _write_sheet(
        wb.create_sheet("Layer 2 — Forecast > SO"), layer2,
        hdr_color=L2_HDR, alt_color=L2_ALT,
        title="Layer 2 — Forecast Exceeds Sales Order  (SO > 0 only)",
        subtitle=f"Grain: SKU × Country × Customer  |  Open months: Jul–Dec 2026  |  Generated: {today_str}  |  {len(layer2):,} records  |  Amber gap = FC > SO",
    )

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary", 0)
    rows = [
        ["SO vs Forecast Mismatch Report"],
        [""],
        ["Generated",    today_str],
        ["Data period",  "Jul – Dec 2026  (open forecast months)"],
        ["Grain",        "SKU × Country × Customer  (one row per combination)"],
        ["Columns",      "SO, FC, Gap per month  (Jul 2026 → Dec 2026)"],
        [""],
        ["Layer", "Description", "Records"],
        ["Layer 1  (Priority)", "Sales Order > AdjFC Forecast (any month)", len(layer1)],
        ["Layer 2", "AdjFC Forecast > Sales Order, SO > 0 (any month)", len(layer2)],
        ["", "Total flagged grains", len(set(layer1.index) | set(layer2.index))],
    ]
    for r_idx, row in enumerate(rows, start=1):
        ws_sum.append(row)
        for cell in ws_sum[r_idx]:
            cell.font   = Font(size=10)
            cell.border = thin

    ws_sum["A1"].font = Font(bold=True, size=14, color=NAVY)
    for cell in ws_sum[8]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE, size=10)
    for cell in ws_sum[9]:
        cell.fill = PatternFill("solid", fgColor=L1_HDR)
        cell.font = Font(bold=True, color=WHITE, size=10)
    for cell in ws_sum[10]:
        cell.fill = PatternFill("solid", fgColor=L2_HDR)
        cell.font = Font(bold=True, color=WHITE, size=10)

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 52
    ws_sum.column_dimensions["C"].width = 14

    wb.save(path)
    print(f"✓ Saved → {path}")


def build_report(output_path: str = None) -> str:
    if output_path is None:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)),
            f"so_fc_mismatch_report_{date.today().isoformat()}.xlsx",
        )

    print("Connecting to BigQuery…")
    client = _client()

    print("Fetching mismatch data…")
    raw = _fetch_data(client)
    print(f"  {len(raw):,} unique grains fetched")

    layer1, layer2 = _build_wide(raw)
    print(f"  Layer 1 (SO > FC): {len(layer1):,} rows")
    print(f"  Layer 2 (FC > SO): {len(layer2):,} rows")

    print("Writing Excel…")
    _write_xlsx(layer1, layer2, output_path)
    return output_path


if __name__ == "__main__":
    path = build_report()
    print(f"\nReport ready: {path}")
