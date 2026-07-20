"""
Email the SO vs Forecast Mismatch Report as an Excel attachment via SendGrid.

Usage:
  python email_report.py          # generate report + send email
  python email_report.py --dry-run  # generate report only, skip sending
"""

import argparse
import base64
import os
import sys
from datetime import date, datetime, timezone, timedelta

import requests

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_so_fc_report import build_report

SENDGRID_API_KEY = os.environ.get("SENDGRID_API_KEY", "")
FROM_EMAIL = "friaz@vpconsulting.mx"
FROM_NAME  = "Aera Demand Planning"

RECIPIENTS = [
    "friaz@vpconsulting.mx",
    "sjensen@vpconsulting.mx",
    "awadood@vpconsulting.mx",
    "emartinez@vpconsulting.mx",
]

PKT = timezone(timedelta(hours=5))


def _html_body(layer1_count: int, layer2_count: int, report_date: str) -> str:
    return f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<style>
  body      {{ font-family: Calibri, Arial, sans-serif; font-size: 14px; color: #1F3864; margin: 0; padding: 0; }}
  .wrapper  {{ max-width: 680px; margin: 0 auto; padding: 32px 24px; }}
  .banner   {{ background: #1F3864; color: #fff; padding: 20px 24px; border-radius: 6px 6px 0 0; }}
  .banner h1{{ margin: 0; font-size: 20px; font-weight: bold; }}
  .banner p {{ margin: 4px 0 0; font-size: 13px; opacity: 0.75; }}
  .body     {{ background: #F9FAFB; padding: 24px; border: 1px solid #D9D9D9; border-top: none; border-radius: 0 0 6px 6px; }}
  .kpi-row  {{ display: flex; gap: 16px; margin: 20px 0; }}
  .kpi      {{ flex: 1; border-radius: 6px; padding: 16px 20px; text-align: center; }}
  .kpi.red  {{ background: #FFF0F0; border: 1px solid #C00000; }}
  .kpi.amb  {{ background: #FFF8E6; border: 1px solid #ED7D31; }}
  .kpi .num {{ font-size: 28px; font-weight: bold; }}
  .kpi.red .num {{ color: #C00000; }}
  .kpi.amb .num {{ color: #ED7D31; }}
  .kpi .lbl {{ font-size: 12px; color: #595959; margin-top: 4px; }}
  .note     {{ font-size: 13px; color: #595959; margin-top: 20px; line-height: 1.6; }}
  .footer   {{ font-size: 11px; color: #999; margin-top: 28px; border-top: 1px solid #E0E0E0; padding-top: 12px; }}
</style>
</head>
<body>
<div class="wrapper">
  <div class="banner">
    <h1>SO vs Forecast Mismatch Report</h1>
    <p>Daily Demand Planning Digest &nbsp;·&nbsp; {report_date}</p>
  </div>
  <div class="body">
    <p>Hi team,</p>
    <p>Please find attached the daily <strong>Sales Order vs Forecast mismatch report</strong> for open months
       <strong>Jul – Dec 2026</strong>. The report flags grains (SKU × Country × Customer) where booked
       sales orders diverge from the adjusted forecast.</p>

    <div class="kpi-row">
      <div class="kpi red">
        <div class="num">{layer1_count:,}</div>
        <div class="lbl">Layer 1 — SO &gt; Forecast<br><small>(Priority flags)</small></div>
      </div>
      <div class="kpi amb">
        <div class="num">{layer2_count:,}</div>
        <div class="lbl">Layer 2 — Forecast &gt; SO<br><small>(SO &gt; 0 only)</small></div>
      </div>
    </div>

    <p class="note">
      <strong>How to read the report:</strong><br>
      • Each row = one SKU × Country × Customer combination<br>
      • Columns show <em>SO</em>, <em>AdjFC</em>, and <em>Gap</em> for each open month<br>
      • <span style="color:#C00000;font-weight:bold;">Red gap</span> = sales order exceeds forecast (Layer 1)<br>
      • <span style="color:#ED7D31;font-weight:bold;">Amber gap</span> = forecast exceeds sales order (Layer 2)
    </p>

    <p class="note">
      Please review Layer 1 first — these are the highest-priority mismatches where booked demand
      exceeds the current forecast plan.
    </p>
  </div>
  <div class="footer">
    Generated automatically by the Aera Demand Planning pipeline &nbsp;·&nbsp; Data refreshed daily
  </div>
</div>
</body>
</html>
""".strip()


def send_email(xlsx_path: str, layer1_count: int, layer2_count: int, dry_run: bool = False):
    report_date = datetime.now(PKT).strftime("%A, %d %b %Y")
    subject     = f"SO vs Forecast Mismatch Report — {report_date}"
    fname       = os.path.basename(xlsx_path)

    if dry_run:
        print(f"[dry-run] Would send '{subject}' to {RECIPIENTS}")
        print(f"[dry-run] Attachment: {fname}")
        return

    if not SENDGRID_API_KEY:
        print("✗ SENDGRID_API_KEY not set — skipping email.")
        return

    with open(xlsx_path, "rb") as f:
        encoded_file = base64.b64encode(f.read()).decode()

    payload = {
        "personalizations": [{"to": [{"email": r} for r in RECIPIENTS]}],
        "from": {"email": FROM_EMAIL, "name": FROM_NAME},
        "subject": subject,
        "content": [{"type": "text/html", "value": _html_body(layer1_count, layer2_count, report_date)}],
        "attachments": [{
            "content": encoded_file,
            "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "filename": fname,
        }],
    }

    print(f"Sending email to {RECIPIENTS}…")
    resp = requests.post(
        "https://api.sendgrid.com/v3/mail/send",
        json=payload,
        headers={"Authorization": f"Bearer {SENDGRID_API_KEY}", "Content-Type": "application/json"},
        timeout=30,
    )
    if resp.status_code == 202:
        print(f"✓ Email sent: '{subject}'")
    else:
        print(f"✗ Email failed: {resp.status_code} — {resp.text}")
        raise RuntimeError(f"SendGrid error {resp.status_code}")


def main(dry_run: bool = False):
    today = date.today().isoformat()
    xlsx_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        f"so_fc_mismatch_report_{today}.xlsx",
    )

    print("Building SO vs Forecast mismatch report…")
    xlsx_path = build_report(output_path=xlsx_path)

    # Count rows per layer from the workbook (avoid re-querying BQ)
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    l1 = wb["Layer 1 — SO > Forecast"].max_row - 4   # subtract title/subtitle/blank/header rows
    l2 = wb["Layer 2 — Forecast > SO"].max_row - 4
    wb.close()

    send_email(xlsx_path, layer1_count=l1, layer2_count=l2, dry_run=dry_run)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Build report but skip sending")
    args = parser.parse_args()
    main(dry_run=args.dry_run)
